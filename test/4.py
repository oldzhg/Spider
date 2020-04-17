from gevent import monkey

monkey.patch_all()
from bs4 import BeautifulSoup
from gevent.queue import Queue
from openpyxl.styles import Font, colors, Alignment
import requests, time, os, openpyxl, gevent

# 爬取薄荷网十一大类食物

start = time.time()

url = 'http://www.boohee.com/food'
res = requests.get(url)
html = res.content.decode('utf-8')
soup = BeautifulSoup(html, 'html.parser')
urls = []

work = Queue()


def find_page():
    for i in soup.find(class_='row').find_all('li'):
        page_urls = []
        food_url = 'http://www.boohee.com' + i.find(class_='img-box').find('a')['href']
        for j in range(1, 11):
            page_url = food_url + '?page=' + str(j)
            page_urls.append(page_url)
        urls.append(page_urls)
    return urls


def get_calory(wb, sheet):
    while not work.empty():
        u = work.get_nowait()
        res = requests.get(u)
        html = res.text
        bs = BeautifulSoup(html, 'html.parser')
        infos = []
        for i in bs.find(class_='food-list').find_all(class_='item clearfix'):
            title = i.find('h4').find('a')['title']
            calory = i.find('p').string
            info = [title, calory]
            infos.append(info)
        sheet['A1'] = '食物名'
        sheet['B1'] = '卡路里'

        title_font = Font(name="等线", size=24, italic=True, color=colors.RED, bold=True)
        content_font = Font(name=u'宋体', size=18)

        center = Alignment(horizontal="center", vertical="center")  # 对齐方式: 数据垂直居中和水平居中

        # font = Font(name=u'宋体', bold=True)
        # align = Alignment(horizontal='center', vertical='center')

        for i in infos:
            sheet.append(i)

        for irow, row in enumerate(sheet.rows, start=1):
            for cell in row:
                cell.font = content_font
                cell.alignment = center

        sheet["A1"].font = sheet["B1"].font = title_font
        sheet["A1"].alignment = sheet["B1"].alignment = center

        wb.save('食物卡路里.xlsx')


def do_work():
    j = 0
    all_urls = find_page()

    for i in range(len(all_urls)):
        for url in all_urls[i]:
            work.put_nowait(url)
        res = requests.get(urls[i][0])
        html = res.text
        bs = BeautifulSoup(html, 'html.parser')
        title = bs.find(class_='widget-food-list pull-right').find('h3').text
        path = './食物卡路里.xlsx'
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            sheet = wb.active  # 进入当前表单
            sheet = wb.create_sheet()
            sheet.title = title
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = title

        task_list = []

        for i in range(5):
            task = gevent.spawn(get_calory(wb, sheet))
            # print("抓取中，请稍后...." + str(j))
            j += 1
            task_list.append(task)
        gevent.joinall(task_list)


do_work()
end = time.time()
interval = end - start
print('耗时%.1f' % interval, 's')
