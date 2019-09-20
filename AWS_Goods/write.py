import re
import requests
from openpyxl import load_workbook

t = 1  # 开始位置
p = 10  # 结束位置  汇总表中数据量共1w
base_url = 'https://www.amazon.ae/s?k=' 
filename = '汇总.xlsx'

wb = load_workbook(filename)  # 打开一个工作簿
sheet = wb['Sheet1']  # 获取表
pattern = re.compile('">AED\s(.*?)</span>')  # 匹配到价格

for i in sheet.iter_rows(min_row=t, max_row=p):  # 从t行到p行迭代
    good_id = i[0].value
    res = requests.get(base_url + str(good_id)) 
    good = re.findall(pattern, res.content.decode())
    if good:  # 如果商品存在
        pos = 'B' + str(t)
        try:
            sheet[pos].value = float(good[0])  # 把价格写入到对应的地方
        except:
            sheet[pos].value = good[0]
        print("已提取 %d " % t)
        wb.save('汇总.xlsx')
    else:
        pos = 'B' + str(t)
        sheet[pos].value = ''  # 价格为空
        print("已提取 %d " % t)
        wb.save('汇总.xlsx')
    t = t + 1

print("表格保存完成！")
